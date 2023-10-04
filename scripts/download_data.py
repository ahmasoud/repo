import psycopg2
from openpyxl import load_workbook

# Define the path to the mapping Excel sheet
mapping_file_path = 'MappingSheet.xlsx'

# Connect to the PostgreSQL database
con = psycopg2.connect(
    database="",
    user="",
    password="",
    host="",
    port=''
)

cursor_obj = con.cursor()

# Read the MasterLine values from the mapping Excel sheet
mapping_workbook = load_workbook(mapping_file_path)
mapping_sheet = mapping_workbook.active

masterline_values = {}

for row in mapping_sheet.iter_rows(min_row=2, values_only=True):
    masterline_id, cell_value = row
    masterline_values[masterline_id] = cell_value

# Execute the SQL query
cursor_obj.execute("SELECT * FROM Updated_Figures.FinalApprovedReport")
result = cursor_obj.fetchall()

# Update the Excel workbook
workbook = load_workbook('8.xlsx')  # Replace '8.xlsx' with the actual file name
sheet = workbook['1']

for row in result:
    masterline_id = row[2]
    if masterline_id in masterline_values:
        cell_value = masterline_values[masterline_id]
        # Assuming the cell address is also defined in the mapping sheet
        cell_address = masterline_values.get(masterline_id)
        if cell_address:
            sheet[cell_address].value = row[4]  # Update the cell value

workbook.save('example_modified.xlsx')

# Close the database connection
con.close()
